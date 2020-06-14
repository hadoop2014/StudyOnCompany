# -*- coding: utf-8 -*-
# @Time    : 12/9/2019 5:03 PM
# @Author  : wu.hao
# @File    : docParserExcel.py
# @Note    : 用于excel文件的读写

from docparser.docParserBaseClass import *
from openpyxl import load_workbook,Workbook
import pandas as pd

class docParserExcel(docParserBase):
    def __init__(self,gConfig,writeParser):
        super(docParserExcel,self).__init__(gConfig)
        self.writeparser = writeParser

    def writeToStore(self, dataFrame, tableName):
        # 专门用于写文件
        # assert self.targetFile != "", "target file %s must not be empty" % self.targetFile
        workbook = load_workbook(self.targetFile)
        if workbook.active.title == "Sheet":  # 表明这是一个空工作薄
            workbook.remove(workbook['Sheet'])  # 删除空工作薄
        writer = pd.ExcelWriter(self.targetFile, engine='openpyxl')
        writer.book = workbook
        dataFrame.to_excel(excel_writer=writer, sheet_name=tableName, index=None)
        # workbook._sheets.insert(0, workbook._sheets.pop())
        writer.save()

    def adjustExcelStyle(self):
        # 调整excel的样式
        # 设置对齐、线性、边框、字体
        from openpyxl.styles import Alignment
        from openpyxl.styles import Side, Border
        from openpyxl.styles import Font

        sheet = self.workbook[self.workbook.sheetnames[0]]
        sheet.insert_rows(idx=0)  # 插入第一行
        font = Font(name='宋体', size=18, bold=True)
        sheet['A1'] = '皮卡丘体育2020年06月新学员信息登记表'
        sheet['A1'].font = font  # 设置字体大小和加粗

        req = ':(\w)'
        weight = re.findall(req, sheet.dimensions)[0]
        sheet.merge_cells(f'A1:{weight}1')

        # 样式先准备好
        alignment = Alignment(horizontal='center', vertical='center')
        side = Side(style='thin', color='000000')
        border = Border(left=side, right=side, top=side, bottom=side)

        # 遍历cell设置样式
        rows = sheet[f'{sheet.dimensions}']
        for row in rows:
            for cell in row:
                cell.alignment = alignment
                cell.border = border

        # 设置前两行的行高
        sheet.row_dimensions[1].height = 38
        sheet.row_dimensions[2].height = 38

        # 设置列宽
        letter_lst = [chr(i + 64).upper() for i in range(2, ord(weight) - ord('A') + 1 + 1)]
        sheet.column_dimensions['A'].width = 8
        for i in letter_lst:
            sheet.column_dimensions[f'{i}'].width = 14

        self.workbook.save(filename=self.targetFile)

    def initialize(self):
        if os.path.exists(self.logging_directory) == False:
            os.makedirs(self.logging_directory)
        if os.path.exists(self.working_directory) == False:
            os.makedirs(self.working_directory)
        self.clear_logging_directory(self.logging_directory)
        suffix = self.targetFile.split('.')[-1]
        assert suffix in self.gConfig['excelSuffix'.lower()], \
            'suffix of %s is invalid,it must one of %s' % (self.targetFile, self.gConfig['excelSuffix'.lower()])
        if self.checkpointIsOn == False:
            # 没有启用备份文件时,初始化workbook
            if os.path.exists(self.targetFile):
                os.remove(self.targetFile)
            self.workbook = Workbook()
            self.writer = pd.ExcelWriter(self.targetFile, engine='openpyxl')
            self.writer.book = self.workbook
            self.writer.save()  # 生成一个新文件

def create_object(gConfig, writeParser=None):
    parser=docParserExcel(gConfig,writeParser)
    parser.initialize()
    return parser