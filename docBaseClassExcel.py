from docBaseClass import  *
from openpyxl import load_workbook,Workbook
import pandas as pd

#深度学习模型的基类
class docBaseExcel(docBase):
    def __init__(self,gConfig):
        super(docBaseExcel,self).__init__(gConfig)

    def writeToExcel(self,dataFrame,sheetName):
        #专门用于写文件
        #assert self.targetFile != "", "target file %s must not be empty" % self.targetFile
        workbook = load_workbook(self.targetFile)
        if workbook.active.title == "Sheet": #表明这是一个空工作薄
            workbook.remove(workbook['Sheet']) #删除空工作薄
        writer = pd.ExcelWriter(self.targetFile, engine='openpyxl')
        writer.book = workbook
        dataFrame.to_excel(excel_writer = writer,sheet_name=sheetName,index=None)
        writer.save()

    def saveCheckpoint(self):
        pass

    def getSaveFile(self):
        if self.model_savefile == '':
            self.model_savefile = None
            return None
        if self.model_savefile is not None:
            if os.path.exists(self.model_savefile)== False:
               return None
                #文件不存在
        return self.model_savefile

    def removeSaveFile(self):
        if self.model_savefile is not None:
            filename = os.path.join(os.getcwd() , self.model_savefile)
            if os.path.exists(filename):
                os.remove(filename)

    def debug_info(self,info = None):
        if self.debugIsOn == False:
            return
        pass
        return

    def debug(self,layer,name=''):
        pass

    def initialize(self):
        if os.path.exists(self.logging_directory) == False:
            os.makedirs(self.logging_directory)
        if os.path.exists(self.working_directory) == False:
            os.makedirs(self.working_directory)
        self.clear_logging_directory(self.logging_directory)
        if os.path.exists(self.targetFile):
            os.remove(self.targetFile)
        suffix = self.targetFile.split('.')[-1]
        assert suffix in self.gConfig['excelSuffix'.lower()],\
            'suffix of %s is invalid,it must one of %s' % (self.targetFile, self.gConfig['excelSuffix'.lower()])
        self.workbook = Workbook()
        self.writer = pd.ExcelWriter(self.targetFile,engine='openpyxl')
        self.writer.book = self.workbook
        self.writer.save()  #生成一个新文件
