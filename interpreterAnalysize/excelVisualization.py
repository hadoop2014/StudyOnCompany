#!/usr/bin/env Python
# coding   : utf-8
# @Time    : 9/6/2020 5:03 PM
# @Author  : wu.hao
# @File    : dataVisualizeation.py
# @Note    : 用于财务数据的可视化

from interpreterAnalysize.interpreterBaseClass import *
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font, colors, Alignment,Border,numbers,PatternFill
from openpyxl.utils import get_column_letter,get_column_interval
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
import pandas as pd


class ExcelVisualization(InterpreterBase):
    def __init__(self,gConfig):
        super(ExcelVisualization, self).__init__(gConfig)
        self.analysizeresult = os.path.join(self.working_directory,gConfig['analysizeresult'])
        self.checkpointIsOn = gConfig['checkpointIsOn'.lower()]


    def _get_class_name(self, gConfig):
        visualization_name = re.findall('(.*)Visualization', self.__class__.__name__).pop().lower()
        return visualization_name


    def read_and_visualize(self,visualize_file,tableName,scale):
        # 专门用于写文件
        visualize_file = os.path.join(self.working_directory,visualize_file)
        #if os.path.exists(visualize_file):
        #    os.remove(visualize_file)
        workbook = Workbook()
        writer = pd.ExcelWriter(visualize_file, engine='openpyxl')
        writer.book = workbook
        #writer.save()  # 生成一个新文件
        workbook = load_workbook(visualize_file)
        writer = pd.ExcelWriter(visualize_file,engine='openpyxl')
        writer.book = workbook
        if workbook.active.title == "Sheet":  # 表明这是一个空工作薄
            workbook.remove(workbook['Sheet'])  # 删除空工作薄

        for reportType in self.gConfig['报告类型']:
            tablePrefix = self._get_tableprefix_by_report_type(reportType)
            sheetName = tablePrefix + tableName# + str(time.thread_time_ns())
            if sheetName in workbook.sheetnames:
                # 先清空旧的工作薄
                workbook.remove(workbook[sheetName])

            dataframe = self._sql_to_dataframe(tableName,sourceTableName=sheetName,scale=scale)

            dataframe = self._process_field_discard(dataframe, tableName)

            self._visualize_to_excel(writer,sheetName,dataframe,tableName)

            self._adjust_style_excel(workbook,sheetName,tableName)
            self.logger.info('%s 展示 %s 成功!' % (scale, sheetName))

        workbook.save(visualize_file)
        workbook.close()


    def _adjust_style_excel(self,workbook,sheetName,tableName):
        sheet = workbook[sheetName]

        self._set_height_and_width(sheet,tableName)

        self._set_freeze_pans(sheet,tableName)

        self._set_font_and_style(sheet,tableName)

        self._set_conditional_formatting(sheet,tableName)

        #blue_fill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
        #dxf = DifferentialStyle(fill=blue_fill, numFmt=NumFmt(10, '0.00%'))
        #rule = Rule('expression', formula=['E3 > 3'], dxf=dxf)
        #ws.conditional_formatting.add('E3', rule)


    def _sql_to_dataframe(self,tableName,sourceTableName,scale):
        if scale == "批量":
            assert ('公司简称' in self.gConfig.keys() and self.gConfig['公司简称'] != NULLSTR) \
                and ('报告时间' in self.gConfig.keys() and self.gConfig['报告时间'] != NULLSTR) \
                and ('报告类型' in self.gConfig.keys() and self.gConfig['报告类型'] != NULLSTR)\
                ,"parameter 公司简称(%s) 报告时间(%s) 报告类型(%s) is not valid parameter"\
                 %(self.gConfig['公司简称'],self.gConfig['报告时间'],self.gConfig['报告类型'])
            #批量处理模式时会进入此分支
            sql = ''
            sql = sql + '\nselect * '
            sql = sql + '\nfrom %s' % (sourceTableName)
            sql = sql + '\nwhere (' + ' or '.join(['公司简称 =' + '\'' +  company + '\''   for company in self.gConfig['公司简称']]) + ')'
            sql = sql + '    and (' + ' or '.join(['报告时间 =' + '\'' + reporttime + '\'' for reporttime in self.gConfig['报告时间']]) + ')'
            sql = sql + '    and (' + ' or '.join(['报告类型 =' + '\'' + reportype + '\'' for reportype in self.gConfig['报告类型']]) + ')'
        else:
            sql = ''
            sql = sql + '\nselect * '
            sql = sql + '\nfrom %s' % (sourceTableName)
        order = self.dictTables[tableName]["order"]
        if isinstance(order,list) and len(order) > 0:
            sql = sql + '\norder by ' + ','.join(order)
        dataframe = pd.read_sql(sql, self._get_connect())
        return dataframe


    def _process_field_discard(self, dataFrame, tableName):
        if dataFrame is None:
            return
        fieldDiscard = self.dictTables[tableName]['fieldDiscard']
        if isinstance(fieldDiscard,list) and len(fieldDiscard) > 0:
            dataFrame = dataFrame.drop(labels=fieldDiscard,axis=1)
        return dataFrame


    def _visualize_to_excel(self,writer,sheetName,dataframe:pd.DataFrame,tableName):
        if dataframe is None:
            return
        startrow = self.dictTables[tableName]['startrow']
        dataframe.to_excel(excel_writer=writer, sheet_name=sheetName, index=None,header=True,startrow=startrow)
        writer.save()


    def _set_conditional_formatting(self,sheet,tableName):
        req = ':([a-zA-Z]+)'
        match = re.findall(req, sheet.dimensions)
        assert isinstance(match, list) and len(match) > 0, 'sheet.dimensions(%s) is invalid!' % sheet.dimensions
        startrow = self.dictTables[tableName]['startrow']
        maxrow = self._get_maxrow(sheet.max_row,tableName)

        conditional_formatting = self.dictTables[tableName]['conditional_formatting']
        conditional_field = conditional_formatting.keys()
        pattern_field = '|'.join(conditional_field)
        weight = match[0]
        letter_list = get_column_interval('A', weight)
        for col_letter in letter_list:
            col = sheet[col_letter]
            field_name = col[startrow].value
            if self._is_matched(pattern_field,field_name):
                operator = conditional_formatting[field_name]['operator']
                if operator != NULLSTR:
                    threhold = conditional_formatting[field_name]['threshold']
                    col[0].value = threhold
                    color_index = conditional_formatting[field_name]['color_index']
                    font = Font(color=colors.COLOR_INDEX[color_index])
                    rule = CellIsRule(operator=operator, formula=[str(threhold)], stopIfTrue=False, font = font)
                    rule_applys = col_letter + str(startrow + 2) + ":" + col_letter + str(maxrow)
                    sheet.conditional_formatting.add(rule_applys, rule)
        return


    def _set_font_and_style(self,sheet, tableName):
        font_settings = self.dictTables[tableName]['font_settings']
        startrow = self.dictTables[tableName]['startrow']
        font_common = Font(name=font_settings['name'], size=font_settings['size'], italic=font_settings['italic']
                    ,color=colors.COLOR_INDEX[font_settings['color_index_common']], bold=font_settings['bold'], underline=None)
        font_emphasize = Font(name=font_settings['name'], size=font_settings['size'], italic=font_settings['italic']
                           , color=colors.COLOR_INDEX[font_settings['color_index_emphasize']], bold=font_settings['bold'],
                           underline=None)
        alignment_header = Alignment(horizontal='left', vertical='center', wrap_text=True)
        alignment_field = Alignment(horizontal='right', vertical='center', wrap_text=False
                                    ,justifyLastLine=True)
        border = Border()
        maxrow = self._get_maxrow(sheet.max_row,tableName)
        freezecol = self.dictTables[tableName]['freezecol']
        builtin_formats = self.dictTables[tableName]['builtin_formats']

        for i, cell in enumerate(sheet[startrow + 1]):
            cell.font = font_common
            cell.alignment = alignment_header
            cell.border = border  # 去掉边框
            if self._is_cell_emphasize(cell, tableName):
                cell.font = font_emphasize
            for index in range(maxrow):
                if self._is_cell_pecentage(cell, tableName):
                    sheet.cell(row=index + 1, column = i + 1).number_format = numbers.FORMAT_PERCENTAGE_00
                elif i + 1 >= freezecol:
                    sheet.cell(row=index + 1, column = i + 1).number_format = numbers.BUILTIN_FORMATS[builtin_formats]
                if index > startrow:
                    sheet.cell(row=index + 1, column = i + 1).alignment = alignment_field


    def _set_freeze_pans(self,sheet,tableName):
        startrow = self.dictTables[tableName]['startrow']
        freezecol = self.dictTables[tableName]['freezecol']
        sheet.freeze_panes = get_column_letter(freezecol) + str(startrow + 2)


    def _set_height_and_width(self,sheet,tableName):
        req = ':([a-zA-Z]+)'
        match = re.findall(req, sheet.dimensions)
        assert isinstance(match,list) and len(match) > 0,'sheet.dimensions(%s) is invalid!'%sheet.dimensions
        #设置标题头的行高
        startrow = self.dictTables[tableName]['startrow']
        maxheight = self.dictTables[tableName]['maxheight']
        sheet.row_dimensions[startrow].height = maxheight
        #设置列宽
        weight = match[0]
        letter_list = get_column_interval('A',weight)
        for col_letter in letter_list:
            widthAdjust = self._get_col_max_width(sheet,col_letter,tableName)
            sheet.column_dimensions[col_letter].width = widthAdjust


    def _get_col_max_width(self,sheet,col_letter,tableName):
        #_get_col_max_width放在_adjust_style_excel的前面跑,这样设置才有效
        maxrow = sheet.max_row
        maxwidth = self.dictTables[tableName]['maxwidth']
        minwidth = self.dictTables[tableName]['minwidth']
        marginwidth = self.dictTables[tableName]['marginwidth']
        startrow = self.dictTables[tableName]['startrow']
        widthAdjust = 0
        col = sheet[col_letter]
        #跳过标题行
        for i in range(startrow + 1,maxrow):
            if len(str(col[i].value)) > widthAdjust:
                widthAdjust = len(str(col[i].value)) + marginwidth
        widthAdjust = min(widthAdjust,maxwidth)
        widthAdjust = max(widthAdjust,minwidth)
        return widthAdjust


    def _column_char_to_integer(self,name):
        assert isinstance(name,str),'the name(%s) of column must be str'%name
        index = 0
        for char in list(name.upper()):
            index = index * 26 + ord(char) - ord('A') + 1
        return index


    def _column_integer_to_char(self,index):
        assert isinstance(index,int),'the index(%s) of column must be integer'%index
        name = NULLSTR
        base = ord('Z') - ord('A') + 1
        while index >= base:
            name = name + chr(index // base + 64).upper()
            index = index - (index // base) * base
        name = name + chr(index % base + 65).upper()
        return name


    def _get_maxrow(self,current_maxrow,tableName):
        maxrow = self.dictTables[tableName]['maxrow']
        return max(maxrow,current_maxrow)


    def _is_cell_pecentage(self,cell,tableName):
        percentage_field = [key for key,value in self.dictTables[tableName]['conditional_formatting'].items()
                            if value['value_format'] == 'percentage']
        pattern_percentage_field = '|'.join(percentage_field)
        isCellPecentage = self._is_matched(pattern_percentage_field, cell.value)
        return isCellPecentage


    def _is_cell_emphasize(self,cell,tableName):
        pattern_emphasize = self.dictTables[tableName]['pattern_emphasize']
        isCellEmphasize = self._is_matched(pattern_emphasize,cell.value)
        return isCellEmphasize


    def initialize(self,dictParameter = None):
        if dictParameter is not None:
            self.gConfig.update(dictParameter)
        if os.path.exists(self.logging_directory) == False:
            os.makedirs(self.logging_directory)
        if os.path.exists(self.working_directory) == False:
            os.makedirs(self.working_directory)


def create_object(gConfig):
    dataVisualization = ExcelVisualization(gConfig)
    dataVisualization.initialize()
    return dataVisualization