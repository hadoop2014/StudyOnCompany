#!/usr/bin/env Python
# coding   : utf-8
# @Time    : 12/9/2019 5:03 PM
# @Author  : wu.hao
# @File    : docParserExcel.py
# @Note    : 用于excel文件的读写

from interpreterAccounting.docparser.docParserBaseClass import *
from openpyxl import load_workbook,Workbook
import pandas as pd

class DocParserExcel(DocParserBase):
    def __init__(self,gConfig):
        super(DocParserExcel, self).__init__(gConfig)
        self.targetFile = os.path.join(self.workingspace.directory,
                                       '.'.join([os.path.split(self.sourceFile)[-1].split('.')[0],'.xlsx']))
        self.workbook = Workbook()

    def writeToStore(self, dictTable):
        # 专门用于写文件
        table = dictTable['table']
        tableName = dictTable['tableName']
        assert len(table) > 0,("%s is empty : %s"%(tableName,dictTable))
        workbook = load_workbook(self.targetFile)
        if workbook.active.title == "Sheet":  # 表明这是一个空工作薄
            workbook.remove(workbook['Sheet'])  # 删除空工作薄
        writer = pd.ExcelWriter(self.targetFile, engine='openpyxl')
        writer.book = workbook
        dataFrame = pd.DataFrame(table, columns=None, index=None)  # 以第一行为列变量
        dataFrame.to_excel(excel_writer=writer, sheet_name=tableName, index=None)
        writer.save()


    def initialize(self,dictParameter=None):
        self.loggingspace.clear_directory(self.loggingspace.directory)
        if dictParameter is not None:
            self.targetFile = os.path.join(self.workingspace.directory,
                                       '.'.join([os.path.split(dictParameter['sourcefile'])[-1].split('.')[0], 'xlsx']))
        suffix = self.targetFile.split('.')[-1]
        assert suffix in self.gConfig['excelSuffix'.lower()], \
            'suffix of %s is invalid,it must one of %s' % (self.targetFile, self.gConfig['excelSuffix'.lower()])
        if os.path.exists(self.targetFile):
            os.remove(self.targetFile)
        self.workbook = Workbook()
        self.writer = pd.ExcelWriter(self.targetFile, engine='openpyxl')
        self.writer.book = self.workbook
        self.writer.save()  # 生成一个新文件

def create_object(gConfig):
    parser=DocParserExcel(gConfig)
    parser.initialize()
    return parser